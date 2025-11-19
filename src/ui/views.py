from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                             QHeaderView, QTableWidgetItem, QFrame, QSizePolicy, QFileDialog)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtCharts import QChart, QChartView, QBarSeries, QBarSet, QBarCategoryAxis, QValueAxis
from PyQt6.QtGui import QPainter
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from qfluentwidgets import (TableWidget, PrimaryPushButton, PushButton, 
                            CalendarPicker, ComboBox, CardWidget, TitleLabel,
                            BodyLabel, StrongBodyLabel, FluentIcon as FIF, InfoBar, InfoBarPosition)

from database.db_manager import DatabaseManager
from .components import TransactionDialog

class DashboardInterface(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("DashboardInterface")
        self.db = DatabaseManager()
        
        self.mainLayout = QHBoxLayout(self)
        self.mainLayout.setContentsMargins(20, 20, 20, 20)
        self.mainLayout.setSpacing(20)

        # --- Left Panel: Filters & Actions ---
        self.leftPanel = CardWidget(self)
        self.leftPanel.setFixedWidth(250)
        self.leftLayout = QVBoxLayout(self.leftPanel)
        
        self.filterLabel = StrongBodyLabel("Фильтры", self.leftPanel)
        self.dateStart = CalendarPicker(self.leftPanel)
        self.dateStart.setDate(QDate.currentDate().addMonths(-1))
        self.dateEnd = CalendarPicker(self.leftPanel)
        self.dateEnd.setDate(QDate.currentDate())
        
        self.applyFilterBtn = PrimaryPushButton("Применить", self.leftPanel)
        self.applyFilterBtn.clicked.connect(self.load_data)
        
        self.resetFilterBtn = PushButton("Сброс", self.leftPanel)
        self.resetFilterBtn.clicked.connect(self.reset_filters)

        self.actionsLabel = StrongBodyLabel("Действия", self.leftPanel)
        self.addBtn = PrimaryPushButton(FIF.ADD, "Добавить", self.leftPanel)
        self.addBtn.clicked.connect(self.show_add_dialog)
        
        self.deleteBtn = PushButton(FIF.DELETE, "Удалить", self.leftPanel)
        self.deleteBtn.clicked.connect(self.delete_transaction)
        
        self.exportExcelBtn = PushButton(FIF.DOCUMENT, "Экспорт Excel", self.leftPanel)
        self.exportExcelBtn.clicked.connect(self.export_to_excel)
        self.exportPdfBtn = PushButton(FIF.PRINT, "Экспорт PDF", self.leftPanel)
        self.exportPdfBtn.clicked.connect(self.export_to_pdf)

        self.leftLayout.addWidget(self.filterLabel)
        self.leftLayout.addWidget(BodyLabel("От:", self.leftPanel))
        self.leftLayout.addWidget(self.dateStart)
        self.leftLayout.addWidget(BodyLabel("До:", self.leftPanel))
        self.leftLayout.addWidget(self.dateEnd)
        self.leftLayout.addWidget(self.applyFilterBtn)
        self.leftLayout.addWidget(self.resetFilterBtn)
        self.leftLayout.addSpacing(20)
        self.leftLayout.addWidget(self.actionsLabel)
        self.leftLayout.addWidget(self.addBtn)
        self.leftLayout.addWidget(self.deleteBtn)
        self.leftLayout.addWidget(self.exportExcelBtn)
        self.leftLayout.addWidget(self.exportPdfBtn)
        self.leftLayout.addStretch(1)

        # --- Center Panel: Table ---
        self.centerPanel = QWidget(self)
        self.centerLayout = QVBoxLayout(self.centerPanel)
        self.centerLayout.setContentsMargins(0, 0, 0, 0)
        
        self.tableTitle = TitleLabel("Транзакции", self.centerPanel)
        self.table = TableWidget(self.centerPanel)
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(['№', 'Дата', 'Тип', 'Категория', 'Сумма', 'Описание'])
        self.table.verticalHeader().hide()
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setEditTriggers(TableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(TableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(TableWidget.SelectionMode.SingleSelection)
        self.table.doubleClicked.connect(self.show_edit_dialog)

        self.centerLayout.addWidget(self.tableTitle)
        self.centerLayout.addWidget(self.table)

        # --- Right Panel: Statistics ---
        self.rightPanel = CardWidget(self)
        self.rightPanel.setFixedWidth(300)
        self.rightLayout = QVBoxLayout(self.rightPanel)
        
        self.statsLabel = StrongBodyLabel("Статистика", self.rightPanel)
        
        # Balance Cards
        self.balanceCard = self.create_stat_card("Баланс", "0.00 ₽")
        self.incomeCard = self.create_stat_card("Доходы", "0.00 ₽", "green")
        self.expenseCard = self.create_stat_card("Расходы", "0.00 ₽", "red")
        
        # Chart
        self.chartView = QChartView(self.rightPanel)
        self.chartView.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.chart = QChart()
        self.chart.setTitle("Доходы и Расходы")
        self.chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        self.chart.legend().setVisible(True)
        self.chart.legend().setAlignment(Qt.AlignmentFlag.AlignBottom)
        self.chartView.setChart(self.chart)
        
        self.rightLayout.addWidget(self.statsLabel)
        self.rightLayout.addWidget(self.balanceCard)
        self.rightLayout.addWidget(self.incomeCard)
        self.rightLayout.addWidget(self.expenseCard)
        self.rightLayout.addWidget(self.chartView, 1)
        self.rightLayout.addStretch(1)

        # Add panels to main layout
        self.mainLayout.addWidget(self.leftPanel)
        self.mainLayout.addWidget(self.centerPanel, 1) # Stretch factor 1
        self.mainLayout.addWidget(self.rightPanel)

        # Initial Data Load
        self.load_data()

    def create_stat_card(self, title, value, color=None):
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 5, 0, 5)
        t = BodyLabel(title, container)
        v = TitleLabel(value, container)
        if color:
            v.setStyleSheet(f"color: {color}")
        layout.addWidget(t)
        layout.addWidget(v)
        return container

    def load_data(self):
        start = self.dateStart.date.toString("yyyy-MM-dd")
        end = self.dateEnd.date.toString("yyyy-MM-dd")
        
        rows = self.db.get_transactions(start, end)
        self.table.setRowCount(len(rows))
        
        for i, row in enumerate(rows):
            # row: (id, date, type, category, amount, desc)
            # We want to display: (i+1, date, type, category, amount, desc)
            
            # Column 0: Sequential Number
            item_no = QTableWidgetItem(str(i + 1))
            item_no.setData(Qt.ItemDataRole.UserRole, row[0]) # Store real ID
            self.table.setItem(i, 0, item_no)
            
            # Other columns
            for j in range(1, 6):
                val = row[j]
                item = QTableWidgetItem(str(val))
                if j == 4: # Amount
                    item.setText(f"{val:.2f}")
                self.table.setItem(i, j, item)
        
        self.update_stats()

    def update_stats(self):
        balance, income, expense = self.db.get_balance()
        self.balanceCard.findChild(TitleLabel).setText(f"{balance:.2f} ₽")
        self.incomeCard.findChild(TitleLabel).setText(f"{income:.2f} ₽")
        self.expenseCard.findChild(TitleLabel).setText(f"{expense:.2f} ₽")
        self.update_chart_data()

    def update_chart_data(self):
        self.chart.removeAllSeries()
        for axis in self.chart.axes():
            self.chart.removeAxis(axis)
            
        summary = self.db.get_monthly_summary()
        
        if not summary:
            return

        set0 = QBarSet("Доходы")
        set1 = QBarSet("Расходы")
        
        categories = sorted(summary.keys())
        # Limit to last 6 months
        categories = categories[-6:]
        
        max_val = 0
        for month in categories:
            inc = summary[month].get('Income', 0.0)
            exp = summary[month].get('Expense', 0.0)
            set0.append(inc)
            set1.append(exp)
            max_val = max(max_val, inc, exp)

        series = QBarSeries()
        series.append(set0)
        series.append(set1)
        self.chart.addSeries(series)

        axisX = QBarCategoryAxis()
        axisX.append(categories)
        self.chart.addAxis(axisX, Qt.AlignmentFlag.AlignBottom)
        series.attachAxis(axisX)

        axisY = QValueAxis()
        axisY.setRange(0, max_val * 1.1 if max_val > 0 else 100)
        self.chart.addAxis(axisY, Qt.AlignmentFlag.AlignLeft)
        series.attachAxis(axisY)

    def reset_filters(self):
        self.dateStart.setDate(QDate.currentDate().addMonths(-1))
        self.dateEnd.setDate(QDate.currentDate())
        self.load_data()

    def export_to_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить Excel", "transactions.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
            
        rows = self.db.get_transactions()
        # Columns: ID, Date, Type, Category, Amount, Description
        # We want to export sequential numbers, not DB IDs
        
        export_data = []
        for i, row in enumerate(rows):
            # row: (id, date, type, category, amount, desc)
            # new row: (i+1, date, type, category, amount, desc)
            new_row = list(row)
            new_row[0] = i + 1
            export_data.append(new_row)

        df = pd.DataFrame(export_data, columns=['№', 'Дата', 'Тип', 'Категория', 'Сумма', 'Описание'])
        
        # Create a Pandas Excel writer using XlsxWriter as the engine
        try:
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Транзакции', index=False, startrow=1)
                
                workbook = writer.book
                worksheet = writer.sheets['Транзакции']
                
                # Define styles
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                centered_alignment = Alignment(horizontal="center", vertical="center")
                border_style = Side(style='thin', color="000000")
                thin_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                
                # Add Title
                worksheet['A1'] = "Отчет по транзакциям"
                worksheet['A1'].font = Font(bold=True, size=16, color="4F81BD")
                worksheet.merge_cells('A1:F1')
                
                # Format Header
                for col_num, value in enumerate(df.columns.values):
                    cell = worksheet.cell(row=2, column=col_num + 1)
                    cell.value = value
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = centered_alignment
                    cell.border = thin_border

                # Format Data
                for row_num in range(3, len(df) + 3):
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="center")
                        
                        # Center Date and Type
                        if col_num in [2, 3]: 
                            cell.alignment = centered_alignment
                        
                        # Format Currency
                        if col_num == 5:
                            cell.number_format = '#,##0.00 ₽'

                # Auto-adjust column widths with buffer
                for i, column_cells in enumerate(worksheet.columns):
                    length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                    # Add extra padding, especially for Amount column (index 4, 0-based)
                    padding = 5
                    if i == 4: # Amount column
                        padding = 10
                    worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + padding

                # --- Add Chart ---
                summary_data = self.db.get_monthly_summary()
                if summary_data:
                    # Create hidden sheet for chart data
                    ws_chart_data = workbook.create_sheet("ChartData")
                    ws_chart_data.sheet_state = 'hidden'
                    
                    headers = ["Месяц", "Доходы", "Расходы"]
                    ws_chart_data.append(headers)
                    
                    months = sorted(summary_data.keys())
                    for month in months:
                        ws_chart_data.append([month, summary_data[month].get('Income', 0), summary_data[month].get('Expense', 0)])
                    
                    chart = BarChart()
                    chart.type = "col"
                    chart.style = 10
                    chart.title = "Динамика Доходов и Расходов"
                    chart.y_axis.title = "Сумма (₽)"
                    chart.x_axis.title = "Месяц"
                    chart.height = 10
                    chart.width = 20
                    
                    data = Reference(ws_chart_data, min_col=2, min_row=1, max_row=len(months)+1, max_col=3)
                    cats = Reference(ws_chart_data, min_col=1, min_row=2, max_row=len(months)+1)
                    
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    
                    worksheet.add_chart(chart, "H2")
            
            InfoBar.success(
                title='Успех',
                content=f'Файл успешно сохранен: {path}',
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP_RIGHT,
                duration=2000,
                parent=self
            )

        except PermissionError:
            InfoBar.error(
                title='Ошибка доступа',
                content=f'Не удалось сохранить файл. Возможно, он открыт в другой программе.\nЗакройте файл "{path}" и попробуйте снова.',
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP_RIGHT,
                duration=5000,
                parent=self
            )
        except Exception as e:
            InfoBar.error(
                title='Ошибка',
                content=f'Произошла ошибка при сохранении: {str(e)}',
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP_RIGHT,
                duration=5000,
                parent=self
            )

    def export_to_pdf(self):
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить PDF", "transactions.pdf", "PDF Files (*.pdf)")
        if not path:
            return

        # Register Cyrillic Font
        try:
            pdfmetrics.registerFont(TTFont('Arial', 'C:\\Windows\\Fonts\\arial.ttf'))
            font_name = 'Arial'
        except:
            font_name = 'Helvetica' # Fallback
            print("Arial font not found, using Helvetica (Cyrillic might not work)")

        doc = SimpleDocTemplate(path, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=font_name,
            fontSize=24,
            spaceAfter=30,
            alignment=1 # Center
        )
        
        # Title
        elements.append(Paragraph("Отчет по финансам", title_style))
        elements.append(Spacer(1, 12))

        # Summary
        balance, income, expense = self.db.get_balance()
        summary_data = [
            [f"Баланс: {balance:.2f} ₽", f"Доходы: {income:.2f} ₽", f"Расходы: {expense:.2f} ₽"]
        ]
        summary_table = Table(summary_data, colWidths=[150, 150, 150])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('TEXTCOLOR', (0, 0), (0, 0), colors.green), # Balance
            ('TEXTCOLOR', (1, 0), (1, 0), colors.blue),  # Income
            ('TEXTCOLOR', (2, 0), (2, 0), colors.red),   # Expense
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # --- Chart ---
        summary = self.db.get_monthly_summary()
        if summary:
            months = sorted(summary.keys())[-6:] # Last 6 months
            incomes = [summary[m].get('Income', 0) for m in months]
            expenses = [summary[m].get('Expense', 0) for m in months]
            
            if months:
                drawing = Drawing(400, 200)
                bc = VerticalBarChart()
                bc.x = 50
                bc.y = 50
                bc.height = 125
                bc.width = 300
                bc.data = [incomes, expenses]
                bc.strokeColor = colors.black
                bc.valueAxis.valueMin = 0
                bc.categoryAxis.labels.boxAnchor = 'ne'
                bc.categoryAxis.labels.dx = 8
                bc.categoryAxis.labels.dy = -2
                bc.categoryAxis.labels.angle = 30
                bc.categoryAxis.categoryNames = months
                bc.bars[0].fillColor = colors.green
                bc.bars[1].fillColor = colors.red
                
                # Legend
                from reportlab.graphics.charts.legends import Legend
                leg = Legend()
                leg.alignment = 'right'
                leg.x = 380
                leg.y = 150
                leg.colorNamePairs = [(colors.green, 'Доходы'), (colors.red, 'Расходы')]
                leg.fontName = font_name
                leg.fontSize = 10
                
                drawing.add(bc)
                drawing.add(leg)
                elements.append(drawing)
                elements.append(Spacer(1, 20))

        # Data Table
        rows = self.db.get_transactions()
        # Header
        data = [['№', 'Дата', 'Тип', 'Категория', 'Сумма', 'Описание']]
        # Rows
        for i, row in enumerate(rows):
            # row: id, date, type, category, amount, desc
            data.append([i + 1, row[1], row[2], row[3], f"{row[4]:.2f}", row[5]])

        table = Table(data, colWidths=[40, 80, 60, 100, 80, 160])
        
        # Table Style
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (4, 1), (4, -1), 'LEFT'), # Description left align
        ])
        table.setStyle(style)
        
        elements.append(table)
        
        # Build
        doc.build(elements)

    def delete_transaction(self):
        row = self.table.currentRow()
        if row < 0:
            InfoBar.warning(
                title='Внимание',
                content='Выберите транзакцию для удаления',
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP_RIGHT,
                duration=2000,
                parent=self
            )
            return

        # Retrieve real ID from UserRole
        t_id = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        
        # Confirm deletion (optional, but good practice)
        # For now, just delete
        try:
            self.db.delete_transaction(t_id)
            self.load_data()
            InfoBar.success(
                title='Успех',
                content='Транзакция удалена',
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP_RIGHT,
                duration=2000,
                parent=self
            )
        except Exception as e:
            InfoBar.error(
                title='Ошибка',
                content=f'Не удалось удалить транзакцию: {e}',
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP_RIGHT,
                duration=2000,
                parent=self
            )

    def show_add_dialog(self):
        dialog = TransactionDialog(self.window())
        if dialog.exec():
            data = dialog.get_data()
            try:
                self.db.add_transaction(**data)
                self.load_data()
            except ValueError as e:
                # Show error (simplified)
                print(f"Error: {e}")

    def show_edit_dialog(self):
        row = self.table.currentRow()
        if row < 0:
            return
            
        # Get data from table
        # Retrieve real ID from UserRole
        t_id = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        date = self.table.item(row, 1).text()
        type_ = self.table.item(row, 2).text()
        cat = self.table.item(row, 3).text()
        amt = float(self.table.item(row, 4).text())
        desc = self.table.item(row, 5).text()
        
        transaction = (t_id, date, type_, cat, amt, desc)
        
        dialog = TransactionDialog(self.window(), transaction)
        if dialog.exec():
            data = dialog.get_data()
            try:
                self.db.update_transaction(t_id, **data)
                self.load_data()
            except ValueError as e:
                print(f"Error: {e}")

